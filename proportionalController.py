from datetime import datetime, timedelta
import time
import openpyxl
from openpyxl import load_workbook

wb = load_workbook('F5.xlsx', data_only=True)
ws = wb.active

DT_INTERVAL = None
DT_TIME = None
TMAX = None
TMIN = None

my_file = open("F6.txt", "w+")
my_file.close()


def parse_f3_file(f3_file):
    with open(f3_file) as file:
        line = file.readline()
        array = line.split(',')
        global DT_INTERVAL, DT_TIME, TMAX, TMIN
        DT_INTERVAL = float(array[0])
        DT_TIME = int(array[1])
        TMAX = int(array[2])
        TMIN = int(array[3])
    new_borders = (TMAX - TMIN) / 4
    TMAX -= new_borders
    TMIN += new_borders


def parse_f5_file(f5_file):
    with open(f5_file) as file:
        line = file.readlines()
    return int(line[-1].split()[-1])


def check_temperature(i):
    # current_t = parse_f5_file('F5.txt')
    current_t = ws.cell(row=i, column=2).value
    if current_t / TMAX > 1:
        with open('F6.txt', 'a+') as f:
            f.write(f"1 {current_t}\n")
        # print(f'1 {current_t}')
    if current_t / TMIN < 1:
        with open('F6.txt', 'a+') as f:
            f.write(f"1 {current_t}\n")
        # print(f'1 {current_t}')


def controller():
    timing = time.time()
    end_time = datetime.now() + timedelta(seconds=DT_TIME)
    i = 2
    while datetime.now() <= end_time:
        if float(time.time() - timing) > float(DT_INTERVAL):
            check_temperature(i)
            i += 1
            timing = time.time()


if __name__ == "__main__":
    parse_f3_file('F3.txt')
    controller()
